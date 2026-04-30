#!/usr/bin/env python3
"""Generate Credit Card Manager Excel file replicating the Angular app's functionality."""

import json
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule

# ── Load data ──────────────────────────────────────────────────────────────────
with open("public/data/cards.json", "r") as f:
    cards = json.load(f)

wb = Workbook()

# ── Style constants ────────────────────────────────────────────────────────────
DARK_BG = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1A1A2E")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="666666", italic=True)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
MONEY_FONT = Font(name="Calibri", size=11, color="1A1A2E")
GREEN_FONT = Font(name="Calibri", bold=True, color="27AE60")
RED_FONT = Font(name="Calibri", bold=True, color="E74C3C")
ORANGE_FONT = Font(name="Calibri", bold=True, color="F39C12")
BLUE_FONT = Font(name="Calibri", bold=True, color="2980B9")

LIGHT_GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

INR_FORMAT = '₹#,##0'
PCT_FORMAT = '0%'

MONTHS_LIST = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def apply_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers, start=col_start):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = DARK_BG
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def is_threshold_met(card):
    return card["spendTillNow"] >= card["spendThreshold"]


def remaining_spend(card):
    r = card["spendThreshold"] - card["spendTillNow"]
    return r if r > 0 else 0


def progress_percent(card):
    return min(card["spendTillNow"] / card["spendThreshold"], 1.0)


def is_urgent(card):
    if is_threshold_met(card):
        return False
    month_name = card["anniversaryMonth"]
    if month_name not in MONTHS_LIST:
        return False
    month_idx = MONTHS_LIST.index(month_name)
    current_month = datetime.now().month - 1  # 0-based
    months_left = month_idx - current_month
    if months_left <= 0:
        months_left += 12
    return months_left <= 3


def get_current_quarter_spend(card):
    now = datetime.now()
    quarter = now.month // 3  # 0-based quarter (roughly)
    # Use same logic as Angular: Math.floor(month / 3)
    current_quarter = math.floor((now.month - 1) / 3)
    quarter_start = current_quarter * 3  # 0-based month
    quarter_month_names = []
    for i in range(3):
        m = quarter_start + i
        d = datetime(now.year, m + 1, 1)
        quarter_month_names.append(d.strftime("%B %Y"))
    total = 0
    for ms in card["monthlySpends"]:
        if ms["month"] in quarter_month_names:
            total += ms["amount"]
    return total


def is_lounge_eligible(card):
    if not card["loungeAccess"]["available"]:
        return False
    return get_current_quarter_spend(card) >= card["loungeAccess"]["quarterlySpendThreshold"]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 1: DASHBOARD
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = "1A1A2E"

# Title
ws_dash.merge_cells("A1:L1")
title_cell = ws_dash["A1"]
title_cell.value = "💳 Credit Card Dashboard"
title_cell.font = TITLE_FONT
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws_dash.row_dimensions[1].height = 35

ws_dash.merge_cells("A2:L2")
ws_dash["A2"].value = "Track your cards, fees & spend thresholds"
ws_dash["A2"].font = SUBTITLE_FONT

# ── Summary Strip (Row 4) ──
total_cards = len(cards)
total_fees = sum(c["annualFee"] for c in cards)
cards_met = sum(1 for c in cards if is_threshold_met(c))

summary_data = [
    ("Total Cards", total_cards),
    ("Total Annual Fees", total_fees),
    ("Threshold Met", f"{cards_met} / {total_cards}"),
]

for i, (label, value) in enumerate(summary_data):
    col = 1 + i * 4
    ws_dash.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 3)
    ws_dash.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 3)
    val_cell = ws_dash.cell(row=4, column=col)
    lbl_cell = ws_dash.cell(row=5, column=col)

    if isinstance(value, int) and label == "Total Annual Fees":
        val_cell.value = value
        val_cell.number_format = INR_FORMAT
    else:
        val_cell.value = value

    val_cell.font = Font(name="Calibri", bold=True, size=18, color="1A1A2E")
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.fill = SUMMARY_FILL
    lbl_cell.value = label
    lbl_cell.font = Font(name="Calibri", size=10, color="666666")
    lbl_cell.alignment = Alignment(horizontal="center")
    lbl_cell.fill = SUMMARY_FILL

    for r in [4, 5]:
        for c in range(col, col + 4):
            ws_dash.cell(row=r, column=c).fill = SUMMARY_FILL
            ws_dash.cell(row=r, column=c).border = THIN_BORDER

# ── Card Table (Row 7+) ──
headers = [
    "Bank Name", "Card Name", "Annual Fee", "Anniversary Month",
    "Spend Threshold", "Spend Till Now", "Remaining", "Progress %",
    "Status", "Insurance", "Lounge Access", "Lounge Eligible"
]
apply_header_row(ws_dash, 7, headers)
ws_dash.row_dimensions[7].height = 25

for idx, card in enumerate(cards):
    row = 8 + idx
    met = is_threshold_met(card)
    urgent = is_urgent(card)
    pct = progress_percent(card)
    rem = remaining_spend(card)

    values = [
        card["bankName"],
        card["cardName"],
        card["annualFee"],
        card["anniversaryMonth"],
        card["spendThreshold"],
        card["spendTillNow"],
        rem,
        pct,
        "✓ Fee Waived" if met else ("⚠ Urgent" if urgent else "In Progress"),
        "🛡️ Included" if card["insurance"]["available"] else "Not Available",
        "Available" if card["loungeAccess"]["available"] else "Not Available",
        ("✓ Eligible" if is_lounge_eligible(card) else "✗ Not Eligible") if card["loungeAccess"]["available"] else "N/A",
    ]

    # Determine row fill
    if met:
        row_fill = LIGHT_GREEN_FILL
    elif urgent:
        row_fill = LIGHT_RED_FILL
    else:
        row_fill = LIGHT_BLUE_FILL if idx % 2 == 0 else None

    for col_idx, val in enumerate(values, start=1):
        cell = ws_dash.cell(row=row, column=col_idx, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if row_fill:
            cell.fill = row_fill

    # Format money columns
    for money_col in [3, 5, 6, 7]:
        ws_dash.cell(row=row, column=money_col).number_format = INR_FORMAT
        ws_dash.cell(row=row, column=money_col).font = MONEY_FONT

    # Format progress
    ws_dash.cell(row=row, column=8).number_format = PCT_FORMAT

    # Status styling
    status_cell = ws_dash.cell(row=row, column=9)
    if met:
        status_cell.font = GREEN_FONT
    elif urgent:
        status_cell.font = RED_FONT
    else:
        status_cell.font = ORANGE_FONT

    # Insurance styling
    ins_cell = ws_dash.cell(row=row, column=10)
    ins_cell.font = GREEN_FONT if card["insurance"]["available"] else Font(name="Calibri", size=11, color="999999")

    # Lounge eligible styling
    lounge_cell = ws_dash.cell(row=row, column=12)
    if card["loungeAccess"]["available"]:
        lounge_cell.font = GREEN_FONT if is_lounge_eligible(card) else RED_FONT

auto_width(ws_dash, min_width=12, max_width=20)
ws_dash.freeze_panes = "A8"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 2: MONTHLY SPENDS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_monthly = wb.create_sheet("Monthly Spends")
ws_monthly.sheet_properties.tabColor = "2980B9"

ws_monthly.merge_cells("A1:H1")
ws_monthly["A1"].value = "📊 Monthly Spend History"
ws_monthly["A1"].font = TITLE_FONT
ws_monthly.row_dimensions[1].height = 35

# Get all unique months
all_months = []
for card in cards:
    for ms in card["monthlySpends"]:
        if ms["month"] not in all_months:
            all_months.append(ms["month"])

headers_monthly = ["Bank Name", "Card Name"] + all_months + ["Total (6 Months)", "Monthly Avg"]
apply_header_row(ws_monthly, 3, headers_monthly)

for idx, card in enumerate(cards):
    row = 4 + idx
    ws_monthly.cell(row=row, column=1, value=card["bankName"]).font = BOLD_FONT
    ws_monthly.cell(row=row, column=1).border = THIN_BORDER
    ws_monthly.cell(row=row, column=2, value=card["cardName"]).font = NORMAL_FONT
    ws_monthly.cell(row=row, column=2).border = THIN_BORDER

    spend_map = {ms["month"]: ms["amount"] for ms in card["monthlySpends"]}
    total_6m = 0
    for mi, month in enumerate(all_months):
        amt = spend_map.get(month, 0)
        total_6m += amt
        cell = ws_monthly.cell(row=row, column=3 + mi, value=amt)
        cell.number_format = INR_FORMAT
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        if idx % 2 == 0:
            cell.fill = LIGHT_BLUE_FILL

    # Total
    total_col = 3 + len(all_months)
    total_cell = ws_monthly.cell(row=row, column=total_col, value=total_6m)
    total_cell.number_format = INR_FORMAT
    total_cell.font = BOLD_FONT
    total_cell.border = THIN_BORDER
    total_cell.fill = LIGHT_GRAY_FILL
    total_cell.alignment = Alignment(horizontal="center")

    # Avg
    avg_cell = ws_monthly.cell(row=row, column=total_col + 1, value=total_6m / max(len(all_months), 1))
    avg_cell.number_format = INR_FORMAT
    avg_cell.font = NORMAL_FONT
    avg_cell.border = THIN_BORDER
    avg_cell.fill = LIGHT_GRAY_FILL
    avg_cell.alignment = Alignment(horizontal="center")

    # Row fill for first 2 cols
    if idx % 2 == 0:
        ws_monthly.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        ws_monthly.cell(row=row, column=2).fill = LIGHT_BLUE_FILL

# Grand total row
grand_row = 4 + len(cards)
ws_monthly.cell(row=grand_row, column=1, value="GRAND TOTAL").font = Font(name="Calibri", bold=True, size=12, color="1A1A2E")
ws_monthly.cell(row=grand_row, column=1).fill = SUMMARY_FILL
ws_monthly.cell(row=grand_row, column=1).border = THIN_BORDER
ws_monthly.cell(row=grand_row, column=2).fill = SUMMARY_FILL
ws_monthly.cell(row=grand_row, column=2).border = THIN_BORDER

grand_total = 0
for mi, month in enumerate(all_months):
    month_total = sum(
        next((ms["amount"] for ms in c["monthlySpends"] if ms["month"] == month), 0)
        for c in cards
    )
    grand_total += month_total
    cell = ws_monthly.cell(row=grand_row, column=3 + mi, value=month_total)
    cell.number_format = INR_FORMAT
    cell.font = BOLD_FONT
    cell.fill = SUMMARY_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")

total_col = 3 + len(all_months)
gt_cell = ws_monthly.cell(row=grand_row, column=total_col, value=grand_total)
gt_cell.number_format = INR_FORMAT
gt_cell.font = Font(name="Calibri", bold=True, size=12, color="1A1A2E")
gt_cell.fill = SUMMARY_FILL
gt_cell.border = THIN_BORDER
gt_cell.alignment = Alignment(horizontal="center")

avg_gt = ws_monthly.cell(row=grand_row, column=total_col + 1, value=grand_total / max(len(all_months), 1))
avg_gt.number_format = INR_FORMAT
avg_gt.font = BOLD_FONT
avg_gt.fill = SUMMARY_FILL
avg_gt.border = THIN_BORDER
avg_gt.alignment = Alignment(horizontal="center")

auto_width(ws_monthly, min_width=14, max_width=22)
ws_monthly.freeze_panes = "C4"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 3: INSURANCE COVERAGE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_ins = wb.create_sheet("Insurance Coverage")
ws_ins.sheet_properties.tabColor = "27AE60"

ws_ins.merge_cells("A1:J1")
ws_ins["A1"].value = "🛡️ Insurance Coverage"
ws_ins["A1"].font = TITLE_FONT
ws_ins.row_dimensions[1].height = 35

ws_ins.merge_cells("A2:J2")
ws_ins["A2"].value = "Cards with insurance benefits"
ws_ins["A2"].font = SUBTITLE_FONT

insurance_cards = [c for c in cards if c["insurance"]["available"]]

# Summary
ws_ins.merge_cells("A3:B3")
ws_ins["A3"].value = f"Insured Cards: {len(insurance_cards)}"
ws_ins["A3"].font = Font(name="Calibri", bold=True, size=13, color="27AE60")

coverage_keys = [
    ("accidentalDeathCover", "💀 Accidental Death Cover"),
    ("airAccidentCover", "✈️ Air Accident Cover"),
    ("fraudLiabilityCover", "🔒 Fraud Liability Cover"),
    ("lostCardLiability", "💳 Lost Card Liability"),
    ("purchaseProtection", "🛍️ Purchase Protection"),
    ("travelInsurance", "🌍 Travel Insurance"),
]

headers_ins = ["Bank Name", "Card Name", "Annual Fee"] + [lbl for _, lbl in coverage_keys] + ["Covers Active"]
apply_header_row(ws_ins, 5, headers_ins)

for idx, card in enumerate(insurance_cards):
    row = 6 + idx
    ws_ins.cell(row=row, column=1, value=card["bankName"]).font = BOLD_FONT
    ws_ins.cell(row=row, column=1).border = THIN_BORDER
    ws_ins.cell(row=row, column=2, value=card["cardName"]).font = NORMAL_FONT
    ws_ins.cell(row=row, column=2).border = THIN_BORDER
    ws_ins.cell(row=row, column=3, value=card["annualFee"]).number_format = INR_FORMAT
    ws_ins.cell(row=row, column=3).font = MONEY_FONT
    ws_ins.cell(row=row, column=3).border = THIN_BORDER

    covers_active = 0
    for ki, (key, _) in enumerate(coverage_keys):
        val = card["insurance"].get(key, 0)
        cell = ws_ins.cell(row=row, column=4 + ki, value=val if val > 0 else "Not Covered")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        if val > 0:
            cell.number_format = INR_FORMAT
            cell.font = Font(name="Calibri", size=11, color="27AE60")
            cell.fill = LIGHT_GREEN_FILL
            covers_active += 1
        else:
            cell.font = Font(name="Calibri", size=11, color="999999")

    # Covers count
    count_cell = ws_ins.cell(row=row, column=4 + len(coverage_keys), value=f"{covers_active}/{len(coverage_keys)}")
    count_cell.font = BOLD_FONT
    count_cell.border = THIN_BORDER
    count_cell.alignment = Alignment(horizontal="center")
    count_cell.fill = LIGHT_GREEN_FILL if covers_active == len(coverage_keys) else LIGHT_ORANGE_FILL

    if idx % 2 == 0:
        for c in [1, 2, 3]:
            ws_ins.cell(row=row, column=c).fill = LIGHT_BLUE_FILL

auto_width(ws_ins, min_width=14, max_width=28)
ws_ins.freeze_panes = "D6"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 4: LOUNGE ACCESS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_lounge = wb.create_sheet("Lounge Access")
ws_lounge.sheet_properties.tabColor = "E67E22"

ws_lounge.merge_cells("A1:J1")
ws_lounge["A1"].value = "✈️ Airport Lounge Cards"
ws_lounge["A1"].font = TITLE_FONT
ws_lounge.row_dimensions[1].height = 35

ws_lounge.merge_cells("A2:J2")
ws_lounge["A2"].value = "Cards with airport lounge facility"
ws_lounge["A2"].font = SUBTITLE_FONT

lounge_cards = [c for c in cards if c["loungeAccess"]["available"]]
eligible_count = sum(1 for c in lounge_cards if is_lounge_eligible(c))

# Summary strip
summary_lounge = [
    ("Lounge Cards", len(lounge_cards)),
    ("Currently Eligible", f"{eligible_count} / {len(lounge_cards)}"),
]

for i, (label, value) in enumerate(summary_lounge):
    col = 1 + i * 4
    ws_lounge.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 3)
    ws_lounge.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 3)
    val_cell = ws_lounge.cell(row=4, column=col)
    val_cell.value = value
    val_cell.font = Font(name="Calibri", bold=True, size=18, color="1A1A2E")
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.fill = SUMMARY_FILL
    lbl_cell = ws_lounge.cell(row=5, column=col)
    lbl_cell.value = label
    lbl_cell.font = Font(name="Calibri", size=10, color="666666")
    lbl_cell.alignment = Alignment(horizontal="center")
    lbl_cell.fill = SUMMARY_FILL
    for r in [4, 5]:
        for c in range(col, col + 4):
            ws_lounge.cell(row=r, column=c).fill = SUMMARY_FILL
            ws_lounge.cell(row=r, column=c).border = THIN_BORDER

headers_lounge = [
    "Bank Name", "Card Name", "Annual Fee", "Anniversary",
    "Quarterly Threshold", "Current Quarter Spend",
    "Progress %", "Remaining", "Status", "Insurance"
]
apply_header_row(ws_lounge, 7, headers_lounge)

for idx, card in enumerate(lounge_cards):
    row = 8 + idx
    eligible = is_lounge_eligible(card)
    qtr_spend = get_current_quarter_spend(card)
    qtr_threshold = card["loungeAccess"]["quarterlySpendThreshold"]
    lounge_pct = min(qtr_spend / qtr_threshold, 1.0) if qtr_threshold > 0 else 0
    lounge_rem = max(qtr_threshold - qtr_spend, 0)

    values = [
        card["bankName"],
        card["cardName"],
        card["annualFee"],
        card["anniversaryMonth"],
        qtr_threshold,
        qtr_spend,
        lounge_pct,
        lounge_rem,
        "✓ Eligible" if eligible else "✗ Not Eligible",
        "🛡️ Included" if card["insurance"]["available"] else "Not Available",
    ]

    row_fill = LIGHT_GREEN_FILL if eligible else (LIGHT_ORANGE_FILL if idx % 2 == 0 else None)

    for ci, val in enumerate(values, start=1):
        cell = ws_lounge.cell(row=row, column=ci, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if row_fill:
            cell.fill = row_fill

    for money_col in [3, 5, 6, 8]:
        ws_lounge.cell(row=row, column=money_col).number_format = INR_FORMAT
    ws_lounge.cell(row=row, column=7).number_format = PCT_FORMAT

    status_cell = ws_lounge.cell(row=row, column=9)
    status_cell.font = GREEN_FONT if eligible else RED_FONT

auto_width(ws_lounge, min_width=14, max_width=24)
ws_lounge.freeze_panes = "A8"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 5: CARD DETAILS & NOTES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_detail = wb.create_sheet("Card Details")
ws_detail.sheet_properties.tabColor = "8E44AD"

ws_detail.merge_cells("A1:F1")
ws_detail["A1"].value = "📝 Card Details & Notes"
ws_detail["A1"].font = TITLE_FONT
ws_detail.row_dimensions[1].height = 35

current_row = 3
for card in cards:
    # Card header
    ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    header_cell = ws_detail.cell(row=current_row, column=1)
    header_cell.value = f"{card['bankName']} — {card['cardName']}"
    header_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    header_cell.fill = DARK_BG
    header_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 7):
        ws_detail.cell(row=current_row, column=c).fill = DARK_BG
        ws_detail.cell(row=current_row, column=c).border = THIN_BORDER
    ws_detail.row_dimensions[current_row].height = 30
    current_row += 1

    # Key info pairs (2 columns of label/value)
    info_pairs = [
        ("Annual Fee", card["annualFee"], INR_FORMAT),
        ("Anniversary Month", card["anniversaryMonth"], None),
        ("Spend Threshold", card["spendThreshold"], INR_FORMAT),
        ("Spend Till Now", card["spendTillNow"], INR_FORMAT),
        ("Remaining", remaining_spend(card), INR_FORMAT),
        ("Progress", f"{progress_percent(card)*100:.0f}%", None),
        ("Status", "✓ Fee Waived" if is_threshold_met(card) else ("⚠ Urgent" if is_urgent(card) else "In Progress"), None),
        ("Insurance", "Active" if card["insurance"]["available"] else "Not Available", None),
        ("Lounge", "Available" if card["loungeAccess"]["available"] else "Not Available", None),
    ]

    for i in range(0, len(info_pairs), 3):
        for j in range(3):
            if i + j < len(info_pairs):
                label, value, fmt = info_pairs[i + j]
                col = 1 + j * 2
                lbl_cell = ws_detail.cell(row=current_row, column=col, value=label)
                lbl_cell.font = Font(name="Calibri", bold=True, size=10, color="666666")
                lbl_cell.border = THIN_BORDER
                lbl_cell.fill = LIGHT_GRAY_FILL
                val_cell = ws_detail.cell(row=current_row, column=col + 1, value=value)
                val_cell.font = BOLD_FONT
                val_cell.border = THIN_BORDER
                if fmt:
                    val_cell.number_format = fmt
        current_row += 1

    # Additional Notes
    if card.get("additionalNotes"):
        ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws_detail.cell(row=current_row, column=1, value="Additional Notes:").font = Font(name="Calibri", bold=True, size=11, color="8E44AD")
        current_row += 1

        notes = card["additionalNotes"].split("\n")
        for note in notes:
            note = note.strip()
            if note:
                ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                note_cell = ws_detail.cell(row=current_row, column=1, value=f"  • {note}")
                note_cell.font = Font(name="Calibri", size=10, color="555555")
                note_cell.alignment = Alignment(wrap_text=True)
                current_row += 1

    current_row += 1  # blank row between cards

auto_width(ws_detail, min_width=15, max_width=35)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 6: SPEND TRACKER (Editable Input Sheet)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_tracker = wb.create_sheet("Spend Tracker")
ws_tracker.sheet_properties.tabColor = "E74C3C"

ws_tracker.merge_cells("A1:H1")
ws_tracker["A1"].value = "📋 Spend Tracker — Update Your Spends Here"
ws_tracker["A1"].font = TITLE_FONT
ws_tracker.row_dimensions[1].height = 35

ws_tracker.merge_cells("A2:H2")
ws_tracker["A2"].value = "Edit the yellow cells to track your monthly spend and threshold progress"
ws_tracker["A2"].font = SUBTITLE_FONT

YELLOW_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

headers_tracker = [
    "Bank Name", "Card Name", "Annual Fee", "Spend Threshold",
    "Spend Till Now (edit →)", "Remaining (auto)", "Progress % (auto)", "Status (auto)"
]
apply_header_row(ws_tracker, 4, headers_tracker)

for idx, card in enumerate(cards):
    row = 5 + idx

    ws_tracker.cell(row=row, column=1, value=card["bankName"]).font = BOLD_FONT
    ws_tracker.cell(row=row, column=1).border = THIN_BORDER
    ws_tracker.cell(row=row, column=2, value=card["cardName"]).font = NORMAL_FONT
    ws_tracker.cell(row=row, column=2).border = THIN_BORDER
    ws_tracker.cell(row=row, column=3, value=card["annualFee"]).number_format = INR_FORMAT
    ws_tracker.cell(row=row, column=3).border = THIN_BORDER
    ws_tracker.cell(row=row, column=4, value=card["spendThreshold"]).number_format = INR_FORMAT
    ws_tracker.cell(row=row, column=4).border = THIN_BORDER

    # Editable cell
    spend_cell = ws_tracker.cell(row=row, column=5, value=card["spendTillNow"])
    spend_cell.number_format = INR_FORMAT
    spend_cell.font = Font(name="Calibri", bold=True, size=12)
    spend_cell.fill = YELLOW_FILL
    spend_cell.border = THIN_BORDER

    # Formula: Remaining = MAX(Threshold - SpendTillNow, 0)
    rem_cell = ws_tracker.cell(row=row, column=6)
    rem_cell.value = f"=MAX(D{row}-E{row},0)"
    rem_cell.number_format = INR_FORMAT
    rem_cell.border = THIN_BORDER
    rem_cell.fill = LIGHT_GRAY_FILL

    # Formula: Progress = MIN(SpendTillNow / Threshold, 1)
    pct_cell = ws_tracker.cell(row=row, column=7)
    pct_cell.value = f"=MIN(E{row}/D{row},1)"
    pct_cell.number_format = PCT_FORMAT
    pct_cell.border = THIN_BORDER
    pct_cell.fill = LIGHT_GRAY_FILL

    # Formula: Status
    status_cell = ws_tracker.cell(row=row, column=8)
    status_cell.value = f'=IF(E{row}>=D{row},"✓ Fee Waived","⏳ In Progress")'
    status_cell.border = THIN_BORDER
    status_cell.fill = LIGHT_GRAY_FILL
    status_cell.font = BOLD_FONT

auto_width(ws_tracker, min_width=14, max_width=26)
ws_tracker.freeze_panes = "A5"

# Add conditional formatting for status column
last_data_row = 4 + len(cards)
ws_tracker.conditional_formatting.add(
    f"H5:H{last_data_row}",
    CellIsRule(operator="equal", formula=['"✓ Fee Waived"'], fill=LIGHT_GREEN_FILL, font=GREEN_FONT)
)
ws_tracker.conditional_formatting.add(
    f"H5:H{last_data_row}",
    CellIsRule(operator="equal", formula=['"⏳ In Progress"'], fill=LIGHT_ORANGE_FILL, font=ORANGE_FONT)
)

# ── Save ───────────────────────────────────────────────────────────────────────
output_path = "Credit_Card_Manager.xlsx"
wb.save(output_path)
print(f"✅ Excel file generated: {output_path}")
print(f"   Sheets: Dashboard | Monthly Spends | Insurance Coverage | Lounge Access | Card Details | Spend Tracker")
